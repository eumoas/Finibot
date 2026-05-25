"""Exportação XLSX do controle financeiro via openpyxl."""
from __future__ import annotations

from collections import defaultdict
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.goal import Goal
from app.models.transaction import Transaction
from app.models.user import User


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INCOME_FILL = PatternFill("solid", fgColor="E2F0D9")
EXPENSE_FILL = PatternFill("solid", fgColor="FCE4D6")


def _money(value: Decimal | float | int) -> float:
    return float(Decimal(str(value or 0)).quantize(Decimal("0.01")))


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(width + 2, 36)


def _summarize(transactions: list[Transaction]) -> dict:
    income = Decimal("0")
    expenses = Decimal("0")
    by_category: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for item in transactions:
        if item.transaction_type == "income":
            income += item.amount
        else:
            expenses += item.amount
            by_category[item.category] += item.amount
    return {
        "income": income,
        "expenses": expenses,
        "balance": income - expenses,
        "by_category": dict(sorted(by_category.items(), key=lambda row: row[1], reverse=True)),
    }


def build_finance_xlsx(
    user: User,
    transactions: list[Transaction],
    goals: list[Goal],
    period_start: date,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lançamentos"

    ws.append(["Data", "Tipo", "Categoria", "Descrição", "Meio de Pagamento", "Valor"])
    for item in sorted(transactions, key=lambda row: row.happened_on):
        row_type = "Receita" if item.transaction_type == "income" else "Despesa"
        ws.append([
            item.happened_on.strftime("%d/%m/%Y"),
            row_type,
            item.category,
            item.description or "",
            "Não informado",
            _money(item.amount),
        ])
        fill = INCOME_FILL if item.transaction_type == "income" else EXPENSE_FILL
        for cell in ws[ws.max_row]:
            cell.fill = fill

    total_row = ws.max_row + 1
    ws.cell(total_row, 5, "Total")
    ws.cell(total_row, 6, f"=SUM(F2:F{max(total_row - 1, 2)})")
    ws.cell(total_row, 5).font = Font(bold=True)
    ws.cell(total_row, 6).font = Font(bold=True)
    _style_header(ws)
    _autosize(ws)

    summary = _summarize(transactions)
    monthly_income = Decimal(str(user.monthly_income or 0))
    ws_summary = wb.create_sheet("Resumo")
    ws_summary.append(["Indicador", "Valor", "% da renda"])
    rows = [
        ("Receitas", summary["income"]),
        ("Despesas", summary["expenses"]),
        ("Saldo", summary["balance"]),
    ]
    for label, value in rows:
        pct = float(value / monthly_income) if monthly_income > 0 else None
        ws_summary.append([label, _money(value), pct])

    ws_summary.append([])
    ws_summary.append(["Categoria", "Total gasto", "% da renda"])
    category_start = ws_summary.max_row + 1
    for category, total in summary["by_category"].items():
        pct = float(total / monthly_income) if monthly_income > 0 else None
        ws_summary.append([category, _money(total), pct])
    category_end = ws_summary.max_row

    _style_header(ws_summary)
    for row in ws_summary.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "0%"

    if category_end >= category_start:
        chart = BarChart()
        chart.title = f"Despesas por categoria - {period_start:%m/%Y}"
        chart.y_axis.title = "Valor"
        chart.x_axis.title = "Categoria"
        data = Reference(ws_summary, min_col=2, min_row=category_start - 1, max_row=category_end)
        cats = Reference(ws_summary, min_col=1, min_row=category_start, max_row=category_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_summary.add_chart(chart, "E2")
    _autosize(ws_summary)

    ws_goals = wb.create_sheet("Metas")
    ws_goals.append(["Meta", "Atual", "Alvo", "Progresso", "Falta", "Prazo"])
    for goal in goals:
        row = ws_goals.max_row + 1
        ws_goals.append([
            goal.title,
            _money(goal.current_amount),
            _money(goal.target_amount),
            f"=IF(C{row}=0,1,B{row}/C{row})",
            f"=MAX(C{row}-B{row},0)",
            goal.deadline.strftime("%d/%m/%Y") if goal.deadline else "",
        ])
    _style_header(ws_goals)
    for row in ws_goals.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = "0%"
    _autosize(ws_goals)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
