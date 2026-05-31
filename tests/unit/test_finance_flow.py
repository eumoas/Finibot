"""Testes unitarios: controle financeiro."""
from datetime import date
from decimal import Decimal
from unittest.mock import AsyncMock, MagicMock
import pytest
from openpyxl import load_workbook
from io import BytesIO
from app.flows.finance_flow import (
    parse_transaction_args,
    parse_transaction_draft,
    _month_range,
    _generate_optional_insight,
    _reply_markdown_safe,
)
from app.flows import finance_flow
from app.flows.xlsx_export import build_finance_xlsx


def test_parse_expense_command_with_decimal_comma():
    parsed = parse_transaction_args("/gasto 18,50 alimentacao lanche escola")

    assert parsed == (Decimal("18.50"), "Alimentação", "lanche escola")


def test_parse_income_command_with_currency_symbol():
    parsed = parse_transaction_args("/receita R$200 mesada")

    assert parsed == (Decimal("200.00"), "Mesada", None)


def test_parse_natural_expense_message():
    parsed = parse_transaction_args("gastei 4.80 no transporte onibus")

    assert parsed == (Decimal("4.80"), "Transporte", "onibus")


def test_parse_informal_transport_message():
    parsed = parse_transaction_args("paguei 12 no busao")

    assert parsed == (Decimal("12.00"), "Transporte", "busao")


def test_parse_natural_income_message_with_preposition():
    parsed = parse_transaction_args("recebi 200 de mesada")

    assert parsed == (Decimal("200.00"), "Mesada", None)


def test_parse_lanche_as_alimentacao():
    parsed = parse_transaction_args("gastei R$18,50 num lanche hoje")

    assert parsed == (Decimal("18.50"), "Alimentação", "lanche")


def test_parse_viagem_category():
    parsed = parse_transaction_args("gastei R$300 em viagem com a galera")

    assert parsed == (Decimal("300.00"), "Viagem", "galera")


def test_parse_informal_viagem_typo_category():
    parsed = parse_transaction_args("gastei 300 na viajem c glr")

    assert parsed == (Decimal("300.00"), "Viagem", "glr")


def test_parse_ontem_sets_previous_date():
    draft = parse_transaction_draft(
        "paguei R$12 no onibus ontem",
        today=date(2026, 5, 20),
    )

    assert draft.transaction_type == "expense"
    assert draft.amount == Decimal("12.00")
    assert draft.category == "Transporte"
    assert draft.happened_on == date(2026, 5, 19)


def test_parse_rejects_missing_amount():
    assert parse_transaction_args("/gasto alimentacao lanche") is None


def test_month_range_handles_december():
    start, end = _month_range(today=date(2026, 12, 20))

    assert start.isoformat() == "2026-12-01"
    assert end.isoformat() == "2027-01-01"


def test_build_finance_xlsx_creates_openpyxl_workbook():
    user = MagicMock()
    user.monthly_income = Decimal("1000.00")

    data = build_finance_xlsx(user, [], [], date(2026, 5, 1))

    workbook = load_workbook(BytesIO(data), data_only=False)
    assert workbook.sheetnames == ["Lançamentos", "Resumo", "Metas"]
    assert workbook["Lançamentos"]["E1"].value == "Meio de Pagamento"


@pytest.mark.asyncio
async def test_optional_insight_omits_technical_error(monkeypatch):
    async def fake_generate_insight(context):
        return "Eita! Tive um problema técnico aqui. Tenta novamente em uns minutinhos? 😅"

    monkeypatch.setattr(finance_flow.llm_gateway, "generate_insight", fake_generate_insight)

    assert await _generate_optional_insight({"saldo": -10}) is None


@pytest.mark.asyncio
async def test_optional_insight_omits_llm_exception(monkeypatch):
    async def fake_generate_insight(context):
        raise RuntimeError("LLM indisponivel")

    monkeypatch.setattr(finance_flow.llm_gateway, "generate_insight", fake_generate_insight)

    assert await _generate_optional_insight({"saldo": -10}) is None


@pytest.mark.asyncio
async def test_reply_markdown_safe_falls_back_to_plain_text():
    message = MagicMock()
    message.reply_text = AsyncMock(side_effect=[RuntimeError("bad markdown"), None])

    await _reply_markdown_safe(message, "Resumo com *meta_ruim*")

    assert message.reply_text.await_count == 2
    first_call, second_call = message.reply_text.await_args_list
    assert first_call.kwargs == {"parse_mode": "Markdown"}
    assert second_call.args == ("Resumo com meta_ruim",)
    assert second_call.kwargs == {}
